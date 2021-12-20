from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from fpdf import FPDF

#file_name = input("veiller enter le nom du fichier, enter the file name: ")
#timeable = input("Veiller enter l'emploi du temps, enter the timeable: ")
#pdfdirectory = input("Veiller scpecifier le dossier ou vous vouler recvoir les convocations")
#if pdfdirectory = virt or pdfdirectory = execel-files:
#   print("impossible")    
#tm = load_worbook(f"excel-files\{timeable})
#wb = load_workbook(f"excel-files\{file_name}") # Selecting students

wb = load_workbook("excel-files\Listes d'Exam S5 EG 2018.2019 pour Etud.xlsx") # Selecting students

ws = wb.active # Selecting the active spreadsheet

#my_range = ws['A'] # Grabs the A column

#for cell in my_range:
#    print(cell.value)

'''
my_row = ws[5]
for cell in my_row:
    print(cell.value)
'''

print(ws[5])
print('\n')
print(ws['A1'].value)
#print(first_row)
# worksheet[row_index_from_1]


# Creatings pdf files
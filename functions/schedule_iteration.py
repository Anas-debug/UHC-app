from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from fpdf import FPDF

fhand = open('table.txt', 'a')

wb = load_workbook("excel-files\Planning d'Examen Automne2018F.xlsx") # Selecting students

ws = wb.active


row = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']

def schedule_iteration(row):
    column = 0
    column_end = len(ws['C'])
    while(column <= column_end):
        column = column + 1
        for x in row:
            exp = f'{x}{column}'
            if ws[exp].value == None :
                continue
            else:
                print(ws[exp].value)

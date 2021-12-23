from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import re

wb = load_workbook("excel-files\Listes d'Exam S5 EG 2018.2019 pour Etud.xlsx") # Selecting students

ws = wb.active

def get_titles(row):
    title_array = []
    column_end = len(ws['A'])
    #print(column_end)
    column = 0
    while(column <= column_end):
        column = column + 1
        for x in row:
            exp = f'{x}{column}'
            if ws[exp].value == None :
             continue
            else:
                if re.search('Liste d\'examen .*', str(ws[exp].value)):
                    title = str(ws[exp].value)
                    title_array.append(ws[exp].value)
                else:
                     continue
                #print(ws[exp].value)
                #arr.append(ws[exp].value)
    #return arr
    print(title_array)




'''
for x in arr:
    if x.find('*Liste'):
        print(x)
'''
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = load_workbook("excel-files\Listes d'Exam S5 EG 2018.2019 pour Etud.xlsx") # Selecting students

ws = wb.active

def enlist_the_values(row):
    arr = []
    column_end = len(ws['A'])
    print(column_end)
    column = 0
    while(column <= column_end):
        column = column + 1
        for x in row:
            exp = f'{x}{column}'
            if ws[exp].value == None :
             continue
            else:
                print(ws[exp].value)
                arr.append(ws[exp].value)
    return arr
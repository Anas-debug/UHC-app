from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from fpdf import FPDF

wb = load_workbook("excel-files\Listes d'Exam S5 EG 2018.2019 pour Etud.xlsx") # Selecting students

ws = wb.active

def iterate_table(row_list):
    column_end = len(ws['A'])
    print(column_end)
    column = 0
    while(column <= column_end):
        column = column + 1
        for x in row_list:
            exp = f'{x}{column}'
            if ws[exp].value == None :
             continue
            else:
                print(ws[exp].value)
